from decimal import Decimal

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse

from openpyxl import Workbook

from core.models import (
    Apresentacao,
    Clinica,
    ImportacaoArquivo,
    Medicamento,
    PerfilUsuario,
)


def criar_xlsx(caminho):
    workbook = Workbook()
    planilha = workbook.active
    planilha.title = "Medicamentos"
    planilha.append(["Medicamento", "Princípio Ativo", "Apresentação", "Concentração", "MG"])
    planilha.append(["Doxorrubicina", "Doxorrubicina", "Frasco 50 mg", "2 mg/mL", "50"])
    planilha.append(["Paclitaxel", "Paclitaxel", "Frasco 300 mg", "6 mg/mL", "300"])
    planilha.append(["", "", "Frasco sem nome", "", "10"])
    planilha.append(["Cisplatina", "Cisplatina", "Frasco 100 mg", "1 mg/mL", "invalido"])
    workbook.save(caminho)


class ImportacaoTests(TestCase):
    def setUp(self):
        self.clinica = Clinica.objects.create(nome="Clínica Importação")
        self.user = get_user_model().objects.create_user(
            username="importadmin", password="Password123456789!"
        )
        PerfilUsuario.objects.create(
            usuario=self.user,
            clinica=self.clinica,
            papel=PerfilUsuario.Papel.ADMINISTRADOR,
            ativo=True,
        )
        self.client.login(username="importadmin", password="Password123456789!")

    def test_inspecionar_importacao_extrai_abas_colunas_e_previa(self):
        from core.services import inspecionar_importacao

        caminho = self.criar_arquivo_temporario()
        abas = inspecionar_importacao(caminho)
        self.assertEqual(abas[0]["nome"], "Medicamentos")
        self.assertEqual(abas[0]["colunas"], ["Medicamento", "Princípio Ativo", "Apresentação", "Concentração", "MG"])
        self.assertEqual(abas[0]["total_linhas"], 4)

    def test_fluxo_upload_preparar_importar(self):
        caminho = self.criar_arquivo_temporario()
        with open(caminho, "rb") as arquivo:
            response = self.client.post(
                reverse("importacoes"),
                {"enviar_arquivo": "1", "arquivo": arquivo},
            )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("importacao_preparar"))

        response = self.client.post(
            reverse("importacao_preparar"),
            {
                "confirmar_importacao": "1",
                "aba": "Medicamentos",
                "map_nome": "0",
                "map_principio_ativo": "1",
                "map_descricao": "2",
                "map_concentracao": "3",
                "map_quantidade_mg": "4",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "2 linha(s) importada(s)")
        medicamento = Medicamento.objects.get(clinica=self.clinica, nome="Doxorrubicina")
        apresentacao = Apresentacao.objects.get(
            medicamento=medicamento, descricao="Frasco 50 mg"
        )
        self.assertEqual(apresentacao.quantidade_mg, Decimal("50"))
        registro = ImportacaoArquivo.objects.get(clinica=self.clinica)
        self.assertEqual(registro.importadas, 2)
        self.assertEqual(registro.com_erro, 2)
        self.client.session.flush()

    def test_importacao_exige_mapeamento_obrigatorio(self):
        caminho = self.criar_arquivo_temporario()
        with open(caminho, "rb") as arquivo:
            self.client.post(
                reverse("importacoes"),
                {"enviar_arquivo": "1", "arquivo": arquivo},
            )
        response = self.client.post(
            reverse("importacao_preparar"),
            {"confirmar_importacao": "1", "aba": "Medicamentos", "map_nome": "0"},
        )
        self.assertEqual(response.status_code, 302)
        self.assertFalse(Medicamento.objects.exists())
        self.client.session.flush()

    def criar_arquivo_temporario(self):
        import os
        import tempfile

        caminho = os.path.join(tempfile.gettempdir(), "teste_importacao_oncologia.xlsx")
        criar_xlsx(caminho)
        return caminho


def criar_xlsx_gmed(caminho):
    workbook = Workbook()
    planilha = workbook.active
    planilha.title = "GMED"
    planilha.append(["NOME DO PRINCÍPIO ATIVO", "CONCENTRAÇÃO", "FORMA FARMACÊUTICA", "NOME DO PRODUTO", "MG"])
    planilha.append(["Doxorrubicina", "2 mg/mL", "Frasco-ampola", "Doxorrubicina 50 mg", "50"])
    planilha.append(["doxorrubicina", "2 mg/mL", "Frasco-ampola", "DOXORRUBICINA 50 MG", "50"])
    planilha.append(["Paclitaxel", "6 mg/mL", "Frasco-ampola", "Paclitaxel 300 mg", "300"])
    planilha.append(["Semmg", "10 mg/mL", "", "Sem concentração em mg", ""])
    workbook.save(caminho)


def criar_xlsx_transferencias(caminho):
    workbook = Workbook()
    planilha = workbook.active
    planilha.title = "Transferencias"
    planilha.append(["Documento", "Data", "Medicamento", "Apresentação", "Qtd", "Lote", "Validade"])
    planilha.append(["TR-JP-2026-001", "01/08/2026", "Doxorrubicina", "Frasco 50 mg", "5", "LOT-DOX-1", "31/12/2026"])
    planilha.append(["TR-JP-2026-001", "01/08/2026", "Paclitaxel", "Frasco 300 mg", "3", "LOT-PAC-1", "31/12/2026"])
    planilha.append(["TR-JP-2026-002", "02/08/2026", "Doxorrubicina", "Frasco 50 mg", "2", "LOT-DOX-2", "31/12/2026"])
    planilha.append(["TR-JP-2026-001", "01/08/2026", "Doxorrubicina", "Frasco 50 mg", "9", "LOT-DOX-X", "31/12/2026"])
    planilha.append(["TR-JP-2026-003", "03/08/2026", "Medicamento Inexistente", "Frasco 10 mg", "1", "", ""])
    workbook.save(caminho)


class ImportacaoGmedTests(TestCase):
    def setUp(self):
        self.clinica = Clinica.objects.create(nome="Clínica GMED")

    def criar_arquivo_temporario(self):
        import os
        import tempfile

        caminho = os.path.join(tempfile.gettempdir(), "teste_importacao_gmed.xlsx")
        criar_xlsx_gmed(caminho)
        return caminho

    def test_gmed_importa_deduplicando_por_nome_normalizado(self):
        from core.services import importar_gmed

        mapeamento = {
            "nome": 3,
            "principio_ativo": 0,
            "descricao": 2,
            "concentracao": 1,
            "quantidade_mg": 4,
        }
        caminho = self.criar_arquivo_temporario()
        importadas, com_erro, erros, novas, duplicadas = importar_gmed(
            self.clinica, caminho, "GMED", mapeamento
        )
        self.assertEqual(importadas, 3)
        self.assertEqual(com_erro, 0)
        self.assertEqual(len(erros), 0)
        self.assertEqual(duplicadas, 1)  # linha 2 = mesma medicação (normalizada)
        self.assertEqual(Medicamento.objects.filter(clinica=self.clinica).count(), 3)
        self.assertEqual(Apresentacao.objects.filter(medicamento__clinica=self.clinica).count(), 2)

    def test_gmed_sem_mg_cria_medicamento_sem_apresentacao(self):
        from core.services import importar_gmed

        mapeamento = {
            "nome": 3,
            "principio_ativo": 0,
            "descricao": 2,
            "concentracao": 1,
            "quantidade_mg": 4,
        }
        caminho = self.criar_arquivo_temporario()
        importadas, _, _, _, duplicadas = importar_gmed(self.clinica, caminho, "GMED", mapeamento)
        self.assertEqual(importadas, 3)
        self.assertEqual(duplicadas, 1)
        sem_mg = Medicamento.objects.get(clinica=self.clinica, principio_ativo="Semmg")
        self.assertEqual(sem_mg.apresentacoes.count(), 0)


class ImportacaoTransferenciasTests(TestCase):
    def setUp(self):
        self.cacoal = Clinica.objects.create(nome="Cacoal")
        self.jiparana = Clinica.objects.create(nome="Ji-Paraná")
        self.medicamento = Medicamento.objects.create(
            clinica=self.cacoal, nome="Doxorrubicina", principio_ativo="Doxorrubicina"
        )
        Apresentacao.objects.create(
            medicamento=self.medicamento,
            concentracao="2 mg/mL",
            descricao="Frasco 50 mg",
            quantidade_mg=Decimal("50"),
        )
        self.paclitaxel = Medicamento.objects.create(
            clinica=self.cacoal, nome="Paclitaxel", principio_ativo="Paclitaxel"
        )
        Apresentacao.objects.create(
            medicamento=self.paclitaxel,
            concentracao="6 mg/mL",
            descricao="Frasco 300 mg",
            quantidade_mg=Decimal("300"),
        )

    def criar_arquivo_temporario(self):
        import os
        import tempfile

        caminho = os.path.join(tempfile.gettempdir(), "teste_importacao_transferencias.xlsx")
        criar_xlsx_transferencias(caminho)
        return caminho

    def test_importa_transferencias_com_dedup_e_ignora_medicamento_desconhecido(self):
        from core.models import Transferencia
        from core.services import importar_transferencias

        mapeamento = {
            "numero": 0,
            "data": 1,
            "medicamento": 2,
            "descricao": 3,
            "quantidade": 4,
            "lote": 5,
            "validade": 6,
        }
        caminho = self.criar_arquivo_temporario()
        importadas, com_erro, erros, total_transferencias = importar_transferencias(
            self.cacoal, self.jiparana, caminho, "Transferencias", mapeamento
        )
        self.assertEqual(importadas, 4)
        self.assertEqual(com_erro, 1)
        self.assertEqual(total_transferencias, 2)
        self.assertTrue(any("não encontrado" in erro for erro in erros))

        transferencia = Transferencia.objects.get(numero="TR-JP-2026-001", clinica_origem=self.jiparana)
        self.assertTrue(transferencia.importada)
        self.assertEqual(transferencia.clinica_destino, self.cacoal)
        self.assertEqual(transferencia.itens.count(), 3)
        self.assertEqual(
            sum(item.quantidade for item in transferencia.itens.all()), 17  # 5 + 3 + 9
        )

    def test_nao_duplica_transferencia_ja_existente(self):
        from core.models import Transferencia
        from core.services import importar_transferencias

        Transferencia.objects.create(
            clinica_origem=self.jiparana,
            clinica_destino=self.cacoal,
            numero="TR-JP-2026-001",
            importada=True,
        )
        mapeamento = {
            "numero": 0,
            "medicamento": 2,
            "descricao": 3,
            "quantidade": 4,
        }
        caminho = self.criar_arquivo_temporario()
        importadas, com_erro, erros, _ = importar_transferencias(
            self.cacoal, self.jiparana, caminho, "Transferencias", mapeamento
        )
        self.assertEqual(importadas, 1)  # apenas TR-JP-2026-002
        self.assertEqual(com_erro, 4)  # 3 linhas duplicadas + medicamento inexistente
        self.assertTrue(any("já existente" in erro for erro in erros))
        self.assertEqual(Transferencia.objects.filter(numero="TR-JP-2026-001").count(), 1)

    def test_fluxo_view_importar_transferencias_e_conciliacao(self):
        from core.models import Transferencia

        user = get_user_model().objects.create_user(
            username="importjp", password="Password123456789!"
        )
        PerfilUsuario.objects.create(
            usuario=user,
            clinica=self.cacoal,
            papel=PerfilUsuario.Papel.ADMINISTRADOR,
            ativo=True,
        )
        self.client.force_login(user)
        caminho = self.criar_arquivo_temporario()
        with open(caminho, "rb") as arquivo:
            response = self.client.post(
                reverse("importacoes"),
                {"enviar_arquivo": "1", "arquivo": arquivo, "importacao_tipo": "transferencias"},
            )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("importacao_preparar"))

        response = self.client.post(
            reverse("importacao_preparar"),
            {
                "confirmar_importacao": "1",
                "aba": "Transferencias",
                "map_numero": "0",
                "map_data": "1",
                "map_medicamento": "2",
                "map_descricao": "3",
                "map_quantidade": "4",
                "map_lote": "5",
                "map_validade": "6",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "4 linha(s) importada(s)")

        response = self.client.get(reverse("importacoes"))
        self.assertContains(response, "Conciliação")
        self.assertContains(response, "TR-JP-2026-001")
        transferencia = Transferencia.objects.get(numero="TR-JP-2026-001")
        self.assertContains(
            response, reverse("detalhe_transferencia", args=[transferencia.pk])
        )
        self.client.session.flush()