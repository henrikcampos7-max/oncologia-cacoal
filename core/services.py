from collections import defaultdict
from decimal import Decimal, ROUND_CEILING
from math import sqrt


def calcular_superficie_corporal(peso_kg, altura_cm):
    if peso_kg is None or altura_cm is None:
        return None
    peso, altura = Decimal(peso_kg), Decimal(altura_cm)
    if peso <= 0 or altura <= 0:
        return None
    return Decimal(str(sqrt(float(peso * altura / Decimal("3600"))))).quantize(
        Decimal("0.01")
    )


def calcular_dose_mg(dose_valor, tipo_dose, peso_kg=None, altura_cm=None):
    dose = Decimal(dose_valor)
    if tipo_dose == "fixa":
        return dose
    if tipo_dose == "mg_kg" and peso_kg is not None:
        return dose * Decimal(peso_kg)
    if tipo_dose == "mg_m2":
        superficie = calcular_superficie_corporal(peso_kg, altura_cm)
        if superficie is not None:
            return dose * superficie
    return None


def calcular_frascos(dose_total_mg, quantidade_mg):
    dose, quantidade = Decimal(dose_total_mg), Decimal(quantidade_mg)
    if dose < 0 or quantidade <= 0:
        raise ValueError("Dose e quantidade por frasco devem ser valores válidos.")
    return int((dose / quantidade).to_integral_value(rounding=ROUND_CEILING))


def numero_na_lista(texto, numero):
    try:
        valores = {int(item.strip()) for item in texto.split(",") if item.strip()}
    except ValueError:
        return False
    return int(numero) in valores


def registrar_auditoria(clinica, usuario, acao, detalhes="", request=None):
    from .models import RegistroAuditoria

    if not clinica:
        return None
    ip_origem = None
    if request is not None:
        ip_origem = request.META.get("HTTP_X_FORWARDED_FOR", "").split(",")[0].strip() or request.META.get("REMOTE_ADDR")
    return RegistroAuditoria.objects.create(
        clinica=clinica,
        usuario=usuario,
        acao=acao,
        detalhes=detalhes,
        ip_origem=ip_origem,
    )


def processar_saida_lotes(clinica, apresentacao, quantidade, usuario=None, observacao=""):
    from .models import Lote, MovimentacaoEstoque

    restante = quantidade
    movimentos = []
    for lote in Lote.objects.filter(
        clinica=clinica,
        apresentacao=apresentacao,
        ativo=True,
        quantidade_atual__gt=0,
    ).order_by("data_validade"):
        if restante <= 0:
            break
        deduzir = min(lote.quantidade_disponivel, restante)
        if deduzir <= 0:
            continue
        lote.quantidade_atual -= deduzir
        lote.save()
        restante -= deduzir
        movimentos.append(
            MovimentacaoEstoque.objects.create(
                clinica=clinica,
                lote=lote,
                tipo=MovimentacaoEstoque.TipoMovimentacao.SAIDA,
                quantidade=-deduzir,
                usuario=usuario,
                observacao=observacao,
            )
        )
    return restante == 0, movimentos


def calcular_estoque_disponivel_apresentacao(clinica, apresentacao):
    from .models import Lote

    return sum(
        lote.quantidade_disponivel
        for lote in Lote.objects.filter(clinica=clinica, apresentacao=apresentacao, ativo=True)
    )


def normalizar_texto(texto):
    """Normaliza texto para deduplicação: caixa baixa, sem acentos, espaços colapsados."""
    import unicodedata

    texto = " ".join(str(texto or "").split())
    texto = texto.casefold()
    return "".join(c for c in unicodedata.normalize("NFD", texto) if unicodedata.category(c) != "Mn")


def inspecionar_importacao(caminho_arquivo, max_linhas_previa=5):
    from openpyxl import load_workbook

    workbook = load_workbook(caminho_arquivo, read_only=True, data_only=True)
    abas = []
    for nome_aba in workbook.sheetnames:
        planilha = workbook[nome_aba]
        linhas = planilha.iter_rows(values_only=True)
        cabecalhos = []
        total = 0
        previa = []
        for indice, linha in enumerate(linhas):
            if indice == 0:
                cabecalhos = [str(valor).strip() if valor is not None else "" for valor in linha]
                continue
            total += 1
            if len(previa) < max_linhas_previa:
                previa.append(list(linha))
        abas.append(
            {
                "nome": nome_aba,
                "colunas": cabecalhos,
                "total_linhas": total,
                "previa": previa,
            }
        )
    workbook.close()
    return abas


def importar_medicamentos(clinica, caminho_arquivo, nome_aba, mapeamento, usuario=None):
    import os
    from decimal import Decimal, InvalidOperation

    from openpyxl import load_workbook

    from .models import Apresentacao, ImportacaoArquivo, Medicamento

    colunas = {indice: campo for campo, indice in mapeamento.items() if campo}
    if not colunas:
        return 0, 0, ["Nenhuma coluna mapeada."], []

    workbook = load_workbook(caminho_arquivo, read_only=True, data_only=True)
    if nome_aba not in workbook.sheetnames:
        workbook.close()
        return 0, 0, [f"Aba '{nome_aba}' não encontrada no arquivo."], []

    planilha = workbook[nome_aba]
    importadas, com_erro = 0, 0
    erros = []
    novas_apresentacoes = []
    for numero, linha in enumerate(planilha.iter_rows(values_only=True), start=1):
        if numero == 1:
            continue
        if not any(valor not in (None, "") for valor in linha):
            continue
        try:
            dados = {}
            for indice, campo in colunas.items():
                dados[campo] = str(linha[indice]).strip() if indice < len(linha) and linha[indice] is not None else ""
            if not dados.get("nome"):
                erros.append(f"Linha {numero}: medicamento sem nome.")
                com_erro += 1
                continue
            if not dados.get("descricao"):
                erros.append(f"Linha {numero} ({dados['nome']}): apresentação sem descrição.")
                com_erro += 1
                continue
            quantidade_mg = dados.get("quantidade_mg") or ""
            try:
                quantidade_mg = Decimal(quantidade_mg.replace(",", "."))
            except InvalidOperation:
                erros.append(
                    f"Linha {numero} ({dados['nome']}): quantidade em mg inválida ('{quantidade_mg}')."
                )
                com_erro += 1
                continue
            if quantidade_mg <= 0:
                erros.append(f"Linha {numero} ({dados['nome']}): quantidade em mg deve ser maior que zero.")
                com_erro += 1
                continue
            medicamento, _ = _buscar_ou_criar_medicamento(
                clinica, dados["nome"].strip(), dados.get("principio_ativo", "").strip()
            )
            apresentacao, criada = _buscar_ou_criar_apresentacao(
                medicamento,
                dados["descricao"].strip(),
                dados.get("concentracao", "").strip(),
                quantidade_mg,
            )
            if criada:
                novas_apresentacoes.append(apresentacao)
            importadas += 1
        except Exception as exc:
            erros.append(f"Linha {numero}: erro inesperado ({exc}).")
            com_erro += 1
    workbook.close()

    _registrar_importacao(
        clinica, caminho_arquivo, nome_aba, importadas, com_erro, erros, usuario,
        tipo=ImportacaoArquivo.Tipo.MEDICAMENTOS,
    )
    return importadas, com_erro, erros, novas_apresentacoes


def importar_gmed(clinica, caminho_arquivo, nome_aba, mapeamento, usuario=None):
    """Importa o catálogo GMED (lista ANVISA): cria/atualiza medicamentos e
    apresentações. Deduplica por nome normalizado (ignora acentos/caixa).

    Campos esperados no mapeamento: nome, principio_ativo, descricao,
    concentracao, quantidade_mg.
    """
    import os
    from decimal import Decimal, InvalidOperation

    from openpyxl import load_workbook

    from .models import ImportacaoArquivo

    colunas = {indice: campo for campo, indice in mapeamento.items() if campo}
    if not colunas:
        return 0, 0, ["Nenhuma coluna mapeada."], []

    workbook = load_workbook(caminho_arquivo, read_only=True, data_only=True)
    if nome_aba not in workbook.sheetnames:
        workbook.close()
        return 0, 0, [f"Aba '{nome_aba}' não encontrada no arquivo."], []

    planilha = workbook[nome_aba]
    importadas, com_erro, duplicadas = 0, 0, 0
    erros = []
    novas_apresentacoes = []
    for numero, linha in enumerate(planilha.iter_rows(values_only=True), start=1):
        if numero == 1:
            continue
        if not any(valor not in (None, "") for valor in linha):
            continue
        try:
            dados = {}
            for indice, campo in colunas.items():
                dados[campo] = str(linha[indice]).strip() if indice < len(linha) and linha[indice] is not None else ""
            if not dados.get("nome"):
                erros.append(f"Linha {numero}: medicamento sem nome.")
                com_erro += 1
                continue
            medicamento, medicamento_criado = _buscar_ou_criar_medicamento(
                clinica, dados["nome"].strip(), dados.get("principio_ativo", "").strip()
            )
            if not medicamento_criado and not dados.get("descricao"):
                duplicadas += 1
                continue
            if not dados.get("descricao"):
                importadas += 1
                continue
            quantidade_mg = dados.get("quantidade_mg") or ""
            if quantidade_mg:
                try:
                    quantidade_mg = Decimal(quantidade_mg.replace(",", "."))
                except InvalidOperation:
                    erros.append(
                        f"Linha {numero} ({dados['nome']}): quantidade em mg inválida ('{quantidade_mg}')."
                    )
                    com_erro += 1
                    continue
                if quantidade_mg <= 0:
                    erros.append(
                        f"Linha {numero} ({dados['nome']}): quantidade em mg deve ser maior que zero."
                    )
                    com_erro += 1
                    continue
            else:
                quantidade_mg = Decimal("1")
            apresentacao, criada = _buscar_ou_criar_apresentacao(
                medicamento,
                dados["descricao"].strip(),
                dados.get("concentracao", "").strip(),
                quantidade_mg,
            )
            if criada:
                novas_apresentacoes.append(apresentacao)
                importadas += 1
            else:
                duplicadas += 1
        except Exception as exc:
            erros.append(f"Linha {numero}: erro inesperado ({exc}).")
            com_erro += 1
    workbook.close()

    _registrar_importacao(
        clinica, caminho_arquivo, nome_aba, importadas, com_erro, erros, usuario,
        tipo=ImportacaoArquivo.Tipo.GMED,
    )
    return importadas, com_erro, erros, novas_apresentacoes, duplicadas


def importar_transferencias(
    clinica_destino,
    clinica_origem,
    caminho_arquivo,
    nome_aba,
    mapeamento,
    usuario=None,
):
    """Importa transferências de Ji-Paraná → Cacoal a partir de planilha.

    Campos esperados no mapeamento: numero, data, medicamento, descricao,
    quantidade, lote, validade.

    Deduplicação: uma transferência (numero) já existente com a mesma origem é
    ignorada e contabilizada como duplicada.
    """
    import os
    from datetime import date, datetime

    from openpyxl import load_workbook

    from .models import ImportacaoArquivo, ItemTransferencia, Transferencia

    colunas = {indice: campo for campo, indice in mapeamento.items() if campo}
    if not colunas:
        return 0, 0, ["Nenhuma coluna mapeada."], 0

    workbook = load_workbook(caminho_arquivo, read_only=True, data_only=True)
    if nome_aba not in workbook.sheetnames:
        workbook.close()
        return 0, 0, [f"Aba '{nome_aba}' não encontrada no arquivo."], 0

    planilha = workbook[nome_aba]
    importadas, com_erro, duplicadas = 0, 0, 0
    erros = []
    transferencias = {}
    for numero, linha in enumerate(planilha.iter_rows(values_only=True), start=1):
        if numero == 1:
            continue
        if not any(valor not in (None, "") for valor in linha):
            continue
        try:
            dados = {}
            for indice, campo in colunas.items():
                dados[campo] = str(linha[indice]).strip() if indice < len(linha) and linha[indice] is not None else ""
            if not dados.get("numero"):
                erros.append(f"Linha {numero}: número do documento ausente.")
                com_erro += 1
                continue
            if not dados.get("medicamento"):
                erros.append(f"Linha {numero} ({dados['numero']}): medicamento ausente.")
                com_erro += 1
                continue
            try:
                quantidade = int(float(dados.get("quantidade", "0")))
            except ValueError:
                quantidade = 0
            if quantidade <= 0:
                erros.append(
                    f"Linha {numero} ({dados['numero']}): quantidade inválida ('{dados.get('quantidade', '')}')."
                )
                com_erro += 1
                continue

            apresentacao = _buscar_apresentacao_por_medicamento(
                clinica_destino, dados["medicamento"], dados.get("descricao", "")
            )
            if apresentacao is None:
                erros.append(
                    f"Linha {numero} ({dados['numero']}): medicamento '{dados['medicamento']}' não encontrado na clínica de destino. Cadastre-o antes de importar."
                )
                com_erro += 1
                continue

            data_documento = None
            if dados.get("data"):
                try:
                    data_documento = date.fromisoformat(dados["data"][:10])
                except ValueError:
                    try:
                        data_documento = datetime.strptime(dados["data"][:10], "%d/%m/%Y").date()
                    except ValueError:
                        erros.append(
                            f"Linha {numero} ({dados['numero']}): data inválida ('{dados['data']}')."
                        )
                        com_erro += 1
                        continue

            transferencia = transferencias.get(dados["numero"])
            if transferencia is None:
                transferencia = Transferencia.objects.filter(
                    clinica_origem=clinica_origem, numero=dados["numero"]
                ).first()
                if transferencia is not None:
                    duplicadas += 1
                    erros.append(
                        f"Linha {numero}: transferência {dados['numero']} já existente (origem {clinica_origem.nome}). Ignorada."
                    )
                    com_erro += 1
                    continue
                transferencia = Transferencia.objects.create(
                    clinica_origem=clinica_origem,
                    clinica_destino=clinica_destino,
                    numero=dados["numero"],
                    importada=True,
                    criado_por=usuario,
                    observacao=f"Importada de planilha (aba '{nome_aba}')",
                )
                transferencias[dados["numero"]] = transferencia

            ItemTransferencia.objects.create(
                transferencia=transferencia,
                apresentacao=apresentacao,
                quantidade=quantidade,
            )
            importadas += 1
        except Exception as exc:
            erros.append(f"Linha {numero}: erro inesperado ({exc}).")
            com_erro += 1
    workbook.close()

    _registrar_importacao(
        clinica_destino, caminho_arquivo, nome_aba, importadas, com_erro, erros, usuario,
        tipo=ImportacaoArquivo.Tipo.TRANSFERENCIAS,
    )
    return importadas, com_erro, erros, len(transferencias)


def _buscar_ou_criar_medicamento(clinica, nome, principio_ativo=""):
    from .models import Medicamento

    chave = normalizar_texto(nome)
    medicamento = None
    for candidato in Medicamento.objects.filter(clinica=clinica):
        if normalizar_texto(candidato.nome) == chave:
            medicamento = candidato
            break
    if medicamento is None:
        medicamento = Medicamento.objects.create(
            clinica=clinica, nome=nome, principio_ativo=principio_ativo
        )
        return medicamento, True
    if principio_ativo and not medicamento.principio_ativo:
        medicamento.principio_ativo = principio_ativo
        medicamento.save(update_fields=["principio_ativo"])
    return medicamento, False


def _buscar_ou_criar_apresentacao(medicamento, descricao, concentracao, quantidade_mg):
    from .models import Apresentacao

    chave = normalizar_texto(descricao)
    for candidato in medicamento.apresentacoes.all():
        if normalizar_texto(candidato.descricao) == chave:
            return candidato, False
    return (
        Apresentacao.objects.create(
            medicamento=medicamento,
            concentracao=concentracao,
            descricao=descricao,
            quantidade_mg=quantidade_mg,
            ativa=True,
        ),
        True,
    )


def _buscar_apresentacao_por_medicamento(clinica, nome_medicamento, descricao=""):
    """Localiza uma apresentação da clínica por nome do medicamento (e descrição opcional)."""
    from .models import Medicamento

    chave_med = normalizar_texto(nome_medicamento)
    for medicamento in Medicamento.objects.filter(clinica=clinica):
        if normalizar_texto(medicamento.nome) != chave_med:
            continue
        if descricao:
            chave_ap = normalizar_texto(descricao)
            for apresentacao in medicamento.apresentacoes.filter(ativa=True):
                if normalizar_texto(apresentacao.descricao) == chave_ap:
                    return apresentacao
            return None
        return medicamento.apresentacoes.filter(ativa=True).first()
    return None


def importar_transferencia_pdf(clinica_origem, clinica_destino, relatorio, usuario=None):
    """Cria transferência a partir do PDF do relatório (Ji-Paraná).

    Fluxo: hash para deduplicação → extração textual → reconhecimento de itens
    via cadastro/aliases → criação da transferência e itens → transição de
    estado para RELATORIO_IMPORTADO. Itens não reconhecidos não bloqueiam a
    importação: viram pendência de conferência manual.
    """
    from decimal import Decimal

    from django.core.files.base import ContentFile

    from .conferencia import transicionar
    from .models import ImportacaoArquivo, ItemTransferencia, Transferencia
    from .relatorio_pdf import calcular_hash, extrair_relatorio, reconhecer_itens

    conteudo = relatorio.read()
    relatorio.seek(0)
    if not conteudo:
        return None, [], ["Arquivo vazio."]

    hash_relatorio = calcular_hash(conteudo)
    if Transferencia.objects.filter(hash_relatorio=hash_relatorio).exists():
        return None, [], ["Relatório já importado anteriormente (hash duplicado)."]

    try:
        dados = extrair_relatorio(conteudo)
    except (ValueError, RuntimeError) as exc:
        return None, [], [f"Falha ao ler relatório: {exc}"]

    # Número legado derivado da referência externa quando existente.
    numero = Transferencia.objects.filter().count() + 1
    exibicao = dados.get("referencia_externa") or f"TR-{numero:04d}"

    transferencia = Transferencia.objects.create(
        clinica_origem=clinica_origem,
        clinica_destino=clinica_destino,
        numero=exibicao,
        importada=True,
        relatorio_arquivo=relatorio,
        hash_relatorio=hash_relatorio,
        data_relatorio=dados.get("data_emissao"),
        referencia_externa=dados.get("referencia_externa", ""),
        criado_por=usuario,
        observacao="Importada de relatório PDF.",
    )
    transferencia.relatorio_arquivo.save(
        f"transferencia_{hash_relatorio[:12]}.pdf",
        ContentFile(conteudo),
        save=True,
    )

    reconhecidos, nao_reconhecidos = reconhecer_itens(clinica_destino, dados["itens"])
    itens_criados = 0
    for item in reconhecidos:
        ItemTransferencia.objects.create(
            transferencia=transferencia,
            apresentacao=item["apresentacao"],
            quantidade=item["quantidade"],
        )
        itens_criados += 1

    erros = list(dados.get("informativo", []))
    for nao in nao_reconhecidos:
        erros.append(f"Item não reconhecido: {nao['descricao'][:60]}")

    _registrar_importacao(
        clinica_destino,
        f"pdf://transferencia/{hash_relatorio[:12]}",
        "relatorio_pdf",
        itens_criados,
        len(erros),
        erros,
        usuario,
        tipo=ImportacaoArquivo.Tipo.TRANSFERENCIAS,
    )
    if usuario is not None:
        transicionar(
            transferencia,
            Transferencia.StatusConferencia.RELATORIO_IMPORTADO,
            usuario=usuario,
            motivo=f"{itens_criados} itens importados do relatório.",
        )
    return transferencia, reconhecidos, erros


def _registrar_importacao(
    clinica, caminho, nome_aba, importadas, com_erro, erros, usuario,
    tipo="",
):
    import os

    from .models import ImportacaoArquivo

    ImportacaoArquivo.objects.create(
        clinica=clinica,
        tipo=tipo,
        nome_arquivo=os.path.basename(caminho),
        aba=nome_aba,
        total_linhas=importadas + com_erro,
        importadas=importadas,
        com_erro=com_erro,
        erros="\n".join(erros[:100]),
        usuario=usuario,
    )


def calcular_sugestao_compras(clinica, dias=30, margem_seguranca=5):
    from datetime import timedelta

    from django.utils import timezone

    from .models import Lote

    hoje = timezone.localdate()
    sessoes_proximas = clinica.sessoes.select_related("paciente", "protocolo").prefetch_related(
        "protocolo__itens__apresentacao__medicamento"
    ).filter(
        data_hora__date__range=(hoje, hoje + timedelta(days=dias)),
        status__in=["agendada", "confirmada"],
    )

    linhas, _ = resumir_sessoes(sessoes_proximas)
    previsao = calcular_previsao_sobras(clinica, dias=dias)
    reaproveitamento = {p["apresentacao_id"]: p for p in previsao["apresentacoes"]}

    sugestoes = []
    for linha in linhas:
        apresentacao = linha["apresentacao_objeto"]
        projecao = reaproveitamento.get(apresentacao.pk)
        frascos_necessarios = projecao["frascos_com_reaproveitamento"] if projecao else linha["frascos"]
        estoque_disponivel = sum(
            l.quantidade_disponivel for l in Lote.objects.filter(
                clinica=clinica, apresentacao=apresentacao, ativo=True
            )
        )
        falta = max(0, frascos_necessarios - estoque_disponivel)
        sugestoes.append(
            {
                "apresentacao": apresentacao,
                "necessario": frascos_necessarios,
                "estoque": estoque_disponivel,
                "falta": falta,
                "sugerido_compra": falta + margem_seguranca if falta > 0 else 0,
                "frascos_sem_reaproveitamento": (
                    projecao["frascos_sem_reaproveitamento"] if projecao else linha["frascos"]
                ),
                "frascos_com_reaproveitamento": frascos_necessarios,
                "economia_frascos": (
                    projecao["economia_frascos"] if projecao else 0
                ),
                "quantidade_reaproveitada_mg": (
                    projecao["quantidade_reaproveitada_mg"] if projecao else Decimal("0")
                ),
                "sobras_iniciais_mg": (
                    projecao["sobras_iniciais_mg"] if projecao else Decimal("0")
                ),
            }
        )
    sugestoes.sort(key=lambda s: (s["apresentacao"].medicamento.nome, s["apresentacao"].descricao))
    return sugestoes


def calcular_previsao_sobras(clinica, dias=30, incluir_sobras_reais=True):
    """Motor de sobras projetadas: processa a agenda futura em ordem cronológica,
    simula abertura de frascos, reaproveita sobras dentro da estabilidade (FEFO),
    calcula perdas projetadas e a necessidade real de frascos por apresentação.

    Não altera estoque físico nem exige conciliação manual. Sobras reais
    disponíveis (quando habilitado) entram no pool como sobras iniciais.
    """
    from datetime import timedelta

    from django.utils import timezone

    sessoes = clinica.sessoes.select_related("paciente", "protocolo").prefetch_related(
        "protocolo__itens__apresentacao__medicamento"
    ).filter(
        data_hora__date__range=(timezone.localdate(), timezone.localdate() + timedelta(days=dias)),
        status__in=["agendada", "confirmada"],
    )

    administracoes = []
    inconsistencias = []
    for sessao in sessoes:
        for item in sessao.protocolo.itens.select_related("apresentacao__medicamento"):
            if not numero_na_lista(item.ciclos, sessao.ciclo):
                continue
            if not numero_na_lista(item.dias_ciclo, sessao.dia_ciclo):
                continue
            dose = calcular_dose_mg(
                item.dose_valor,
                item.tipo_dose,
                sessao.paciente.peso_kg,
                sessao.paciente.altura_cm,
            )
            if dose is None:
                inconsistencias.append(
                    f"Dados insuficientes para calcular {item.apresentacao} na sessão {sessao.pk}."
                )
                continue
            administracoes.append(
                {
                    "sessao_id": sessao.pk,
                    "paciente": sessao.paciente,
                    "data_hora": sessao.data_hora,
                    "apresentacao": item.apresentacao,
                    "dose_mg": dose,
                }
            )

    administracoes.sort(key=lambda adm: (adm["data_hora"], adm["sessao_id"]))

    sobras_iniciais_por_apresentacao = {}
    if incluir_sobras_reais:
        for sobra in sobras_reais_validas(clinica):
            sobras_iniciais_por_apresentacao.setdefault(sobra.apresentacao_id, []).append(
                {
                    "restante_mg": sobra.quantidade_mg,
                    "limite": sobra.limite_estabilidade,
                    "origem": (
                        sobra.paciente_origem.nome if sobra.paciente_origem else "Sobra real"
                    ),
                }
            )

    por_apresentacao = defaultdict(list)
    for adm in administracoes:
        por_apresentacao[adm["apresentacao"].pk].append(adm)

    resultados = []
    for ap_id, lista in sorted(
        por_apresentacao.items(),
        key=lambda item: (
            item[1][0]["apresentacao"].medicamento.nome,
            item[1][0]["apresentacao"].descricao,
        ),
    ):
        sobras_iniciais = sobras_iniciais_por_apresentacao.get(ap_id)
        resultados.append(
            _simular_reaproveitamento(lista[0]["apresentacao"], lista, sobras_iniciais)
        )

    return {"apresentacoes": resultados, "inconsistencias": inconsistencias}


def sobras_reais_validas(clinica):
    """Sobras reais disponíveis e ainda dentro da estabilidade."""
    from django.utils import timezone

    from .models import SobraReal

    return list(
        SobraReal.objects.select_related("apresentacao", "paciente_origem").filter(
            clinica=clinica,
            status=SobraReal.Status.DISPONIVEL,
            limite_estabilidade__gte=timezone.now(),
        )
    )


def sobras_reais_expiradas(clinica):
    """Sobras reais disponíveis cuja estabilidade já terminou (marca para expirar)."""
    from django.utils import timezone

    from .models import SobraReal

    return SobraReal.objects.filter(
        clinica=clinica,
        status=SobraReal.Status.DISPONIVEL,
        limite_estabilidade__lt=timezone.now(),
    )


def _simular_reaproveitamento(apresentacao, administracoes, sobras_iniciais=None):
    """Simula aberturas e reaproveitamentos para uma única apresentação.

    `sobras_iniciais` (opcional) permite iniciar o pool com sobras projetadas
    pré-existentes — ponto de integração futuro para sobras reais no motor.

    Retorna indicadores e a linha do tempo (eventos) da projeção.
    """
    quantidade_mg = apresentacao.quantidade_mg
    tem_estabilidade = apresentacao.estabilidade_cadastrada
    sobras = [dict(s) for s in (sobras_iniciais or [])]  # {"restante_mg", "limite", "origem"}
    eventos = []
    pacientes = set()
    administracoes_total = 0
    demanda_total = Decimal("0")
    frascos_sem = 0
    frascos_com = 0
    reaproveitada = Decimal("0")
    perda_projetada = Decimal("0")

    for adm in administracoes:
        data_hora, dose = adm["data_hora"], adm["dose_mg"]
        nome_paciente = adm["paciente"].nome
        pacientes.add(nome_paciente)
        administracoes_total += 1
        demanda_total += dose
        frascos_sem += calcular_frascos(dose, quantidade_mg)

        if tem_estabilidade:
            expiradas = [s for s in sobras if data_hora > s["limite"]]
            for s in expiradas:
                perda_projetada += s["restante_mg"]
                eventos.append(
                    {
                        "tipo": "perda_projetada",
                        "paciente": s["origem"],
                        "data_hora": s["limite"],
                        "mg": s["restante_mg"],
                        "detalhe": "estabilidade vencida",
                    }
                )
            sobras = [s for s in sobras if data_hora <= s["limite"]]

            sobras.sort(key=lambda s: s["limite"])
            restante = dose
            for s in sobras:
                if restante <= 0:
                    break
                usar = min(s["restante_mg"], restante)
                s["restante_mg"] -= usar
                restante -= usar
                reaproveitada += usar
                eventos.append(
                    {
                        "tipo": "reuso",
                        "paciente": nome_paciente,
                        "data_hora": data_hora,
                        "mg": usar,
                        "detalhe": f"sobra de {s['origem']}",
                    }
                )
            sobras = [s for s in sobras if s["restante_mg"] > 0]
            dose_restante = restante
        else:
            dose_restante = dose

        if dose_restante > 0:
            frascos = calcular_frascos(dose_restante, quantidade_mg)
            frascos_com += frascos
            eventos.append(
                {
                    "tipo": "abertura",
                    "paciente": nome_paciente,
                    "data_hora": data_hora,
                    "mg": dose_restante,
                    "detalhe": f"{frascos} frasco(s) de {quantidade_mg} mg",
                }
            )
            sobra_mg = frascos * quantidade_mg - dose_restante
            if tem_estabilidade and sobra_mg > 0:
                limite = apresentacao.limite_estabilidade_desde(data_hora)
                sobras.append(
                    {"restante_mg": sobra_mg, "limite": limite, "origem": nome_paciente}
                )

    return {
        "apresentacao_id": apresentacao.pk,
        "apresentacao_objeto": apresentacao,
        "medicamento": apresentacao.medicamento.nome,
        "apresentacao": apresentacao.descricao,
        "estabilidade_cadastrada": tem_estabilidade,
        "flag": None if tem_estabilidade else "ESTABILIDADE_NAO_CADASTRADA",
        "pacientes": sorted(pacientes),
        "quantidade_pacientes": len(pacientes),
        "administracoes": administracoes_total,
        "demanda_total_mg": demanda_total.quantize(Decimal("0.01")),
        "frascos_sem_reaproveitamento": frascos_sem,
        "frascos_com_reaproveitamento": frascos_com,
        "quantidade_reaproveitada_mg": reaproveitada.quantize(Decimal("0.01")),
        "perda_projetada_mg": perda_projetada.quantize(Decimal("0.01")),
        "economia_frascos": max(0, frascos_sem - frascos_com),
        "sobras_iniciais_mg": sum(
            (s["restante_mg"] for s in (sobras_iniciais or [])), Decimal("0")
        ).quantize(Decimal("0.01")),
        "frascos_necessarios": frascos_com,
        "eventos": eventos,
    }


def resumir_sessoes(sessoes):
    acumulado = defaultdict(lambda: {"administracoes": 0, "dose_total": Decimal("0")})
    inconsistencias = []
    for sessao in sessoes:
        for item in sessao.protocolo.itens.select_related("apresentacao__medicamento"):
            if not numero_na_lista(item.ciclos, sessao.ciclo):
                continue
            if not numero_na_lista(item.dias_ciclo, sessao.dia_ciclo):
                continue
            dose = calcular_dose_mg(
                item.dose_valor,
                item.tipo_dose,
                sessao.paciente.peso_kg,
                sessao.paciente.altura_cm,
            )
            if dose is None:
                inconsistencias.append(
                    f"Dados insuficientes para calcular {item.apresentacao} na sessão {sessao.pk}."
                )
                continue
            chave = item.apresentacao_id
            acumulado[chave]["apresentacao"] = item.apresentacao
            acumulado[chave]["administracoes"] += 1
            acumulado[chave]["dose_total"] += dose

    linhas = []
    for dados in acumulado.values():
        apresentacao = dados["apresentacao"]
        linhas.append(
            {
                "medicamento": apresentacao.medicamento.nome,
                "apresentacao": apresentacao.descricao,
                "apresentacao_objeto": apresentacao,
                "administracoes": dados["administracoes"],
                "dose_total": dados["dose_total"].quantize(Decimal("0.01")),
                "quantidade_mg": apresentacao.quantidade_mg,
                "frascos": calcular_frascos(dados["dose_total"], apresentacao.quantidade_mg),
            }
        )
    linhas.sort(key=lambda linha: (linha["medicamento"], linha["apresentacao"]))
    return linhas, inconsistencias


def processar_baixa_estoque_sessao(sessao, usuario=None):
    """
    Processa a baixa no estoque por FEFO (First Expired, First Out) para cada item do protocolo da sessão realizada.
    Quando há estabilidade cadastrada, a sobra física resultante (mg) é registrada
    automaticamente como SobraReal no pool de reaproveitamento.
    """
    from .models import Lote, MovimentacaoEstoque, SobraReal
    from django.utils import timezone

    if sessao.movimentacoes_estoque.exists():
        return True, ["Baixa de estoque já havia sido realizada para esta sessão."]

    mensagens = []
    from django.db import transaction

    with transaction.atomic():
        for item in sessao.protocolo.itens.select_related("apresentacao"):
            if not numero_na_lista(item.ciclos, sessao.ciclo):
                continue
            if not numero_na_lista(item.dias_ciclo, sessao.dia_ciclo):
                continue

            dose = calcular_dose_mg(
                item.dose_valor,
                item.tipo_dose,
                sessao.paciente.peso_kg,
                sessao.paciente.altura_cm,
            )
            if dose is None or dose == 0:
                continue

            frascos_necessarios = calcular_frascos(dose, item.apresentacao.quantidade_mg)
            if frascos_necessarios <= 0:
                continue

            lotes_disponiveis = Lote.objects.filter(
                clinica=sessao.clinica,
                apresentacao=item.apresentacao,
                ativo=True,
                quantidade_atual__gt=0,
            ).order_by("data_validade")

            restante = frascos_necessarios
            lote_utilizado = None
            for lote in lotes_disponiveis:
                if restante <= 0:
                    break
                deduzir = min(lote.quantidade_disponivel, restante)
                if deduzir <= 0:
                    continue
                lote.quantidade_atual -= deduzir
                lote.save()
                lote_utilizado = lote
                restante -= deduzir

                MovimentacaoEstoque.objects.create(
                    clinica=sessao.clinica,
                    lote=lote,
                    tipo=MovimentacaoEstoque.TipoMovimentacao.SAIDA,
                    quantidade=-deduzir,
                    sessao=sessao,
                    usuario=usuario,
                    observacao=f"Aplicação realizada — Paciente: {sessao.paciente.nome}",
                )

            if restante > 0:
                mensagens.append(
                    f"Estoque insuficiente para {item.apresentacao.descricao}. Faltaram {restante} frascos."
                )
                continue

            sobra_mg = frascos_necessarios * item.apresentacao.quantidade_mg - dose
            if sobra_mg > 0 and item.apresentacao.estabilidade_cadastrada:
                limite = item.apresentacao.limite_estabilidade_desde(sessao.data_hora)
                if limite is not None and limite > timezone.now():
                    SobraReal.objects.create(
                        clinica=sessao.clinica,
                        apresentacao=item.apresentacao,
                        quantidade_mg=sobra_mg,
                        lote=lote_utilizado,
                        paciente_origem=sessao.paciente,
                        data_abertura=sessao.data_hora,
                        limite_estabilidade=limite,
                        criada_por=usuario,
                    )
                    mensagens.append(
                        f"Sobra de {('%.3f' % sobra_mg).rstrip('0').rstrip('.')} mg registrada para {item.apresentacao.descricao}."
                    )

    return True, mensagens


def coletar_alertas_estoque(clinica):
    """Retorna (vencidos, validade_critica, validade_alerta, estoque_baixo) da clínica."""
    lotes = clinica.lotes.filter(ativo=True).select_related("apresentacao__medicamento")
    vencidos, criticos_validade, alertas_validade, estoque_baixo = [], [], [], []
    for lote in lotes:
        validade, estoque = lote.status_validade, lote.status_estoque
        if validade == "vencido":
            vencidos.append(lote)
        elif validade == "critico":
            criticos_validade.append(lote)
        elif validade == "alerta":
            alertas_validade.append(lote)
        if estoque in ("baixo", "esgotado"):
            estoque_baixo.append(lote)
    return vencidos, criticos_validade, alertas_validade, estoque_baixo


def coletar_faltas_recentes(clinica, limite_dias=30, minimo_faltas=2):
    """Pacientes com faltas recorrentes no período, por ordem de quantidade."""
    from datetime import timedelta

    from django.db.models import Count
    from django.utils import timezone

    from .models import SessaoTratamento

    desde = timezone.localdate() - timedelta(days=limite_dias)
    return list(
        SessaoTratamento.objects.filter(
            clinica=clinica,
            status=SessaoTratamento.Status.FALTOU,
            data_hora__date__gte=desde,
        )
        .values("paciente__id", "paciente__nome")
        .annotate(total=Count("id"))
        .filter(total__gte=minimo_faltas)
        .order_by("-total")
    )


def enviar_alertas_por_email(clinica, usuario=None, request=None):
    """Envia um resumo dos alertas da clínica para administradores e farmacêuticos.

    Retorna (quantidade_de_destinatarios, total_de_alertas).
    """
    from django.core.mail import send_mail

    from .models import PerfilUsuario

    vencidos, criticos_validade, alertas_validade, estoque_baixo = coletar_alertas_estoque(clinica)
    faltas = coletar_faltas_recentes(clinica)
    total_alertas = len(vencidos) + len(criticos_validade) + len(estoque_baixo) + len(faltas)
    if total_alertas == 0:
        return 0, 0

    destinatarios = list(
        PerfilUsuario.objects.filter(
            clinica=clinica,
            ativo=True,
            papel__in=[PerfilUsuario.Papel.ADMINISTRADOR, PerfilUsuario.Papel.FARMACEUTICO],
        )
        .exclude(usuario__email="")
        .values_list("usuario__email", flat=True)
        .distinct()
    )
    if not destinatarios:
        return 0, total_alertas

    linhas = []
    if vencidos:
        linhas.append(f"Lotes VENCIDOS ({len(vencidos)}):")
        linhas += [
            f"- {lote.apresentacao} — Lote {lote.numero_lote} (validade {lote.data_validade:%d/%m/%Y})"
            for lote in vencidos[:10]
        ]
    if criticos_validade:
        linhas.append(f"Validade crítica em até 30 dias ({len(criticos_validade)}):")
        linhas += [
            f"- {lote.apresentacao} — Lote {lote.numero_lote} (validade {lote.data_validade:%d/%m/%Y})"
            for lote in criticos_validade[:10]
        ]
    if estoque_baixo:
        linhas.append(f"Estoque baixo ou esgotado ({len(estoque_baixo)}):")
        linhas += [
            f"- {lote.apresentacao} — Lote {lote.numero_lote} (atual: {lote.quantidade_atual} frascos)"
            for lote in estoque_baixo[:10]
        ]
    if faltas:
        linhas.append(f"Pacientes com faltas recorrentes nos últimos 30 dias ({len(faltas)}):")
        linhas += [f"- {item['paciente__nome']} ({item['total']} falta(s))" for item in faltas]

    corpo = f"Resumo de alertas — {clinica.nome}\n\n" + "\n".join(linhas)
    send_mail(
        subject=f"[Oncologia Cacoal] {total_alertas} alerta(s) em {clinica.nome}",
        message=corpo,
        from_email=None,
        recipient_list=destinatarios,
        fail_silently=True,
    )
    registrar_auditoria(
        clinica,
        usuario,
        "Envio de alertas por email",
        f"Enviado para {len(destinatarios)} destinatário(s) com {total_alertas} alerta(s).",
        request=request,
    )
    return len(destinatarios), total_alertas

